import pandas as pd
import numpy as np
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.axis import DateAxis
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler
import warnings
warnings.filterwarnings('ignore')

def safe_round(value, default=0):
    try:
        if pd.isna(value) or value is None:
            return default
        return int(round(float(value)))
    except:
        return default

def load_and_preprocess_data():
    print("\n📂 PASSO 1: Carregar arquivo CSV de dados históricos")
    print("Por favor, forneça o arquivo CSV com as colunas: DT_VENDA, DS_PRODUTO, CD_PRODUTO, UND_MED, QTD_SAIDA")
    filepath = input("Digite o caminho completo do arquivo CSV ou arraste o arquivo para aqui: ").strip('"')

    try:
        df = pd.read_csv(filepath, sep=",", parse_dates=["DT_VENDA"], dayfirst=True)
    except Exception as e:
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                content = f.read()
            content = content.replace(',"', '||"').replace(',', '.').replace('||"', ',"')
            with open("temp_file.csv", "w", encoding="utf-8") as f:
                f.write(content)
            df = pd.read_csv("temp_file.csv", sep=",", parse_dates=["DT_VENDA"], dayfirst=True)
            os.remove("temp_file.csv")
        except:
            print(f"❌ Erro ao ler o arquivo: {e}")
            return None

    required_cols = ['DT_VENDA', 'DS_PRODUTO', 'CD_PRODUTO', 'UND_MED', 'QTD_SAIDA']
    if not all(col in df.columns for col in required_cols):
        print("❌ O arquivo CSV não contém todas as colunas necessárias")
        print(f"Colunas necessárias: {required_cols}")
        print(f"Colunas encontradas: {df.columns.tolist()}")
        return None

    df['QTD_SAIDA'] = df['QTD_SAIDA'].apply(lambda x: safe_round(x))
    df.columns = df.columns.str.strip()
    df["DT_VENDA"] = pd.to_datetime(df["DT_VENDA"], errors='coerce', dayfirst=True)
    df = df.dropna(subset=["DT_VENDA"])

    q1 = df["QTD_SAIDA"].quantile(0.05)
    q3 = df["QTD_SAIDA"].quantile(0.95)
    df = df[(df["QTD_SAIDA"] >= q1) & (df["QTD_SAIDA"] <= q3)]

    print("✅ Dados históricos carregados e pré-processados com sucesso")
    return df

def load_real_data():
    print("\n📂 PASSO 2: Carregar arquivo CSV de dados reais de 2025")
    print("Por favor, forneça o arquivo CSV com as colunas: ANOMES_VENDA, CD_PRODUTO, DS_PRODUTO, QTD_SAIDA")
    filepath = input("Digite o caminho completo do arquivo CSV de dados reais: ").strip('"')

    try:
        df_real = pd.read_csv(filepath, sep=",")
        
        # Converter ANOMES_VENDA para data
        df_real['ANOMES_VENDA'] = df_real['ANOMES_VENDA'].astype(str)
        df_real['ano'] = df_real['ANOMES_VENDA'].str[:4]
        df_real['mes'] = df_real['ANOMES_VENDA'].str[4:6]
        df_real['data'] = pd.to_datetime(df_real['ano'] + '-' + df_real['mes'] + '-01')
        
        # Limpar e processar dados
        df_real['QTD_SAIDA'] = df_real['QTD_SAIDA'].apply(lambda x: safe_round(x))
        df_real = df_real[['data', 'CD_PRODUTO', 'DS_PRODUTO', 'QTD_SAIDA']]
        
        print("✅ Dados reais carregados com sucesso")
        return df_real
        
    except Exception as e:
        print(f"❌ Erro ao ler o arquivo de dados reais: {e}")
        return None

def feature_engineering(df):
    print("\n🔧 Criando features para análise...")
    df_grouped = df.groupby(["DT_VENDA", "CD_PRODUTO", "DS_PRODUTO", "UND_MED"])["QTD_SAIDA"].sum().reset_index()
    produtos = df_grouped["CD_PRODUTO"].unique()
    full_dfs = []

    for produto in produtos:
        df_prod = df_grouped[df_grouped["CD_PRODUTO"] == produto].copy()
        df_prod = df_prod.sort_values("DT_VENDA")

        date_range = pd.date_range(start=df_prod["DT_VENDA"].min(), end=df_prod["DT_VENDA"].max(), freq='D')
        df_full = pd.DataFrame({'DT_VENDA': date_range})
        df_full = df_full.merge(df_prod, on='DT_VENDA', how='left')
        df_full["CD_PRODUTO"] = df_full["CD_PRODUTO"].fillna(produto)
        df_full["DS_PRODUTO"] = df_full["DS_PRODUTO"].fillna(df_prod["DS_PRODUTO"].iloc[0])
        df_full["UND_MED"] = df_full["UND_MED"].fillna(df_prod["UND_MED"].iloc[0])
        df_full["QTD_SAIDA"] = df_full["QTD_SAIDA"].fillna(0)

        df_full['dia_semana'] = df_full['DT_VENDA'].dt.dayofweek
        df_full['mes'] = df_full['DT_VENDA'].dt.month
        df_full['trimestre'] = df_full['DT_VENDA'].dt.quarter
        df_full['final_semana'] = df_full['dia_semana'].isin([5, 6]).astype(int)

        df_full['media_movel_7d'] = df_full['QTD_SAIDA'].rolling(window=7, min_periods=1).mean()
        df_full['media_movel_30d'] = df_full['QTD_SAIDA'].rolling(window=30, min_periods=1).mean()
        df_full = df_full.fillna(method='bfill')

        full_dfs.append(df_full)

    return pd.concat(full_dfs)

def train_and_predict(df_full):
    print("\n🤖 Treinando modelos de Machine Learning...")
    produtos = df_full["CD_PRODUTO"].unique()
    all_forecasts = []

    for produto in produtos:
        df_prod = df_full[df_full["CD_PRODUTO"] == produto].copy()
        df_prod = df_prod.sort_values("DT_VENDA")

        X = df_prod[['dia_semana', 'mes', 'trimestre', 'final_semana', 'media_movel_7d', 'media_movel_30d']]
        y = df_prod["QTD_SAIDA"]

        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, shuffle=False)

        scaler = StandardScaler()
        X_train_scaled = scaler.fit_transform(X_train)
        X_test_scaled = scaler.transform(X_test)

        models = {
            'Random Forest': RandomForestRegressor(n_estimators=100, random_state=42),
            'Linear Regression': LinearRegression()
        }

        best_model = None
        best_score = -np.inf

        for name, model in models.items():
            model.fit(X_train_scaled, y_train)
            score = model.score(X_test_scaled, y_test)
            if score > best_score:
                best_score = score
                best_model = model

        datas_prev = pd.date_range(start="2025-01-01", end="2025-12-31", freq="D")
        future_data = pd.DataFrame({'DT_VENDA': datas_prev})
        future_data['dia_semana'] = future_data['DT_VENDA'].dt.dayofweek
        future_data['mes'] = future_data['DT_VENDA'].dt.month
        future_data['trimestre'] = future_data['DT_VENDA'].dt.quarter
        future_data['final_semana'] = future_data['dia_semana'].isin([5, 6]).astype(int)
        future_data['media_movel_7d'] = df_prod['media_movel_7d'].iloc[-1]
        future_data['media_movel_30d'] = df_prod['media_movel_30d'].iloc[-1]

        X_future = future_data[['dia_semana', 'mes', 'trimestre', 'final_semana', 'media_movel_7d', 'media_movel_30d']]
        X_future_scaled = scaler.transform(X_future)

        y_future = best_model.predict(X_future_scaled)
        y_future = np.clip(y_future, 0, None).round().astype(int)

        df_forecast = pd.DataFrame({
            'data': datas_prev,
            'qtd_prevista': y_future,
            'CD_PRODUTO': produto,
            'DS_PRODUTO': df_prod['DS_PRODUTO'].iloc[0],
            'UND_MED': df_prod['UND_MED'].iloc[0]
        })

        all_forecasts.append(df_forecast)

    return pd.concat(all_forecasts)

def calculate_inventory_levels(demand_series):
    total_demand = safe_round(demand_series.sum())

    estoque_minimo = total_demand * 1.05
    estoque_ideal = total_demand * 1.15
    estoque_maximo = total_demand * 1.25

    return {
        'estoque_minimo': safe_round(estoque_minimo),
        'estoque_ideal': safe_round(estoque_ideal),
        'estoque_maximo': safe_round(estoque_maximo)
    }

def create_comparison_analysis(df_forecasts, df_real):
    print("\n📊 Criando análise comparativa...")
    
    # Agregar previsões por mês
    df_monthly_forecast = df_forecasts.copy()
    df_monthly_forecast["mes_ano"] = df_monthly_forecast["data"].dt.to_period("M")
    monthly_forecast = df_monthly_forecast.groupby(["mes_ano", "CD_PRODUTO", "DS_PRODUTO", "UND_MED"])["qtd_prevista"].sum().reset_index()
    
    # Agregar dados reais por mês
    df_real["mes_ano"] = df_real["data"].dt.to_period("M")
    monthly_real = df_real.groupby(["mes_ano", "CD_PRODUTO", "DS_PRODUTO"])["QTD_SAIDA"].sum().reset_index()
    
    # Fazer merge dos dados
    comparison = monthly_forecast.merge(
        monthly_real, 
        left_on=["mes_ano", "CD_PRODUTO"], 
        right_on=["mes_ano", "CD_PRODUTO"], 
        how="inner",
        suffixes=("_prev", "_real")
    )
    
    # Calcular níveis de estoque para cada mês
    comparison_with_levels = []
    
    for _, row in comparison.iterrows():
        levels = calculate_inventory_levels(pd.Series([row['qtd_prevista']]))
        
        # Calcular percentual de erro
        if row['QTD_SAIDA'] != 0:
            erro_percentual = ((row['qtd_prevista'] - row['QTD_SAIDA']) / row['QTD_SAIDA']) * 100
        else:
            erro_percentual = 0 if row['qtd_prevista'] == 0 else 100
        
        # Diferenças
        diff_demanda = row['qtd_prevista'] - row['QTD_SAIDA']
        diff_minimo = levels['estoque_minimo'] - row['QTD_SAIDA']
        diff_ideal = levels['estoque_ideal'] - row['QTD_SAIDA']
        diff_maximo = levels['estoque_maximo'] - row['QTD_SAIDA']
        
        comparison_with_levels.append({
            'Mês': row['mes_ano'].strftime('%Y-%m'),
            'Código Produto': row['CD_PRODUTO'],
            'Produto': row['DS_PRODUTO_prev'],
            'Unidade': row['UND_MED'],
            'Demanda Real': safe_round(row['QTD_SAIDA']),
            'Demanda Prevista': safe_round(row['qtd_prevista']),
            'Estoque Mínimo': levels['estoque_minimo'],
            'Estoque Ideal': levels['estoque_ideal'],
            'Estoque Máximo': levels['estoque_maximo'],
            'Erro %': round(erro_percentual, 2),
            'Diff Demanda': diff_demanda,
            'Diff Mínimo': diff_minimo,
            'Diff Ideal': diff_ideal,
            'Diff Máximo': diff_maximo
        })
    
    return pd.DataFrame(comparison_with_levels)

def create_excel_charts(wb, df_comparison):
    print("\n📈 Criando gráficos dos 10 principais produtos no Excel...")
    
    # Selecionar os 10 produtos com maior demanda real total
    top_products = df_comparison.groupby(['Código Produto', 'Produto'])['Demanda Real'].sum().nlargest(10)
    
    chart_sheet_num = 1
    for (cod_produto, nome_produto), _ in top_products.items():
        df_produto = df_comparison[df_comparison['Código Produto'] == cod_produto].copy()
        df_produto = df_produto.sort_values('Mês')
        
        # Criar nova aba para cada produto
        ws_chart = wb.create_sheet(f"Gráfico_{chart_sheet_num}")
        
        # Adicionar dados do produto na planilha
        headers = ['Mês', 'Demanda Real', 'Demanda Prevista', 'Estoque Mínimo', 'Estoque Ideal', 'Estoque Máximo', 'Erro %']
        ws_chart.append(headers)
        
        for _, row in df_produto.iterrows():
            ws_chart.append([
                row['Mês'],
                row['Demanda Real'],
                row['Demanda Prevista'],
                row['Estoque Mínimo'],
                row['Estoque Ideal'],
                row['Estoque Máximo'],
                row['Erro %']
            ])
        
        # Criar gráfico de linha para demanda
        line_chart = LineChart()
        line_chart.title = f"{nome_produto[:30]} - Comparativo de Demanda"
        line_chart.style = 13
        line_chart.y_axis.title = 'Quantidade'
        line_chart.x_axis.title = 'Mês'
        
        data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(df_produto)+1, max_col=6)
        cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(df_produto)+1)
        
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(cats)
        
        ws_chart.add_chart(line_chart, "I2")
                
        # Ajustar largura das colunas
        for col in ws_chart.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws_chart.column_dimensions[col[0].column_letter].width = max(10, min(30, max_length + 2))
        
        chart_sheet_num += 1
        if chart_sheet_num > 10:  # Limitar a 10 produtos
            break

def generate_excel_report(df_forecasts, df_comparison, output_file="previsao_comparativo_2025.xlsx"):
    print("\n📊 Gerando relatório Excel...")
    wb = Workbook()

    # Aba 1: Previsão Mensal
    ws_monthly = wb.create_sheet("Previsão Mensal")
    df_monthly = df_forecasts.copy()
    df_monthly["mes"] = df_monthly["data"].dt.to_period("M").dt.to_timestamp()
    monthly_grouped = df_monthly.groupby(["mes", "CD_PRODUTO", "DS_PRODUTO", "UND_MED"])["qtd_prevista"]

    data = []
    for (mes, produto, nome, unidade), group in monthly_grouped:
        levels = calculate_inventory_levels(group)
        data.append({
            'Mês': mes.strftime('%Y-%m'),
            'Código Produto': produto,
            'Produto': nome,
            'Unidade': unidade,
            'Demanda Prevista': safe_round(group.sum()),
            'Estoque Mínimo': levels['estoque_minimo'],
            'Estoque Ideal': levels['estoque_ideal'],
            'Estoque Máximo': levels['estoque_maximo']
        })

    df_mensal = pd.DataFrame(data)
    for r in dataframe_to_rows(df_mensal, index=False, header=True):
        ws_monthly.append(r)

    # Aba 2: Comparativo Real vs Previsto
    ws_comparison = wb.create_sheet("Comparativo Real vs Previsto")
    for r in dataframe_to_rows(df_comparison, index=False, header=True):
        ws_comparison.append(r)

    # Criar gráficos dos 10 principais produtos
    create_excel_charts(wb, df_comparison)

    # Formatação das colunas
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(10, min(50, max_length + 2))

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_file)
    print(f"✅ Arquivo Excel salvo: {output_file}")

def create_summary_report(df_comparison):
    print("\n📋 Criando resumo da análise...")
    
    # Resumo geral
    total_produtos = df_comparison['Código Produto'].nunique()
    total_comparacoes = len(df_comparison)
    erro_medio = df_comparison['Erro %'].mean()
    erro_absoluto_medio = df_comparison['Erro %'].abs().mean()
    
    # Top 10 produtos com maior demanda real
    top_products = df_comparison.groupby(['Código Produto', 'Produto'])['Demanda Real'].sum().nlargest(10)
    
    # Produtos com maior erro
    produtos_maior_erro = df_comparison.groupby(['Código Produto', 'Produto'])['Erro %'].mean().abs().nlargest(5)
    
    print(f"""
    ================================
    RESUMO DA ANÁLISE COMPARATIVA
    ================================
    
    📊 ESTATÍSTICAS GERAIS:
    • Total de produtos analisados: {total_produtos}
    • Total de comparações (produto/mês): {total_comparacoes}
    • Erro médio: {erro_medio:.2f}%
    • Erro absoluto médio: {erro_absoluto_medio:.2f}%
    
    🏆 TOP 10 PRODUTOS POR DEMANDA REAL:
    """)
    
    for i, ((cod, nome), demanda) in enumerate(top_products.items(), 1):
        print(f"    {i:2d}. {cod} - {nome[:40]}... (Demanda: {demanda:,})")
    
    print(f"""
    ⚠️  TOP 5 PRODUTOS COM MAIOR ERRO MÉDIO:
    """)
    
    for i, ((cod, nome), erro) in enumerate(produtos_maior_erro.items(), 1):
        print(f"    {i}. {cod} - {nome[:40]}... (Erro: {erro:.2f}%)")

def main():
    print("""
    ================================
    PREVISÃO DE DEMANDA - 2025
    COM COMPARATIVO DE DADOS REAIS
    ================================
    """)

    # Carregar dados históricos
    df = load_and_preprocess_data()
    if df is None:
        return

    # Carregar dados reais de 2025
    df_real = load_real_data()
    if df_real is None:
        return

    # Processar dados e fazer previsões
    df_full = feature_engineering(df)
    df_forecasts = train_and_predict(df_full)

    # Criar análise comparativa
    df_comparison = create_comparison_analysis(df_forecasts, df_real)

    # Gerar relatório Excel com gráficos
    generate_excel_report(df_forecasts, df_comparison)

    # Mostrar resumo da análise
    create_summary_report(df_comparison)

    print("\n🏁 Processo concluído com sucesso!")
    print("📁 Arquivo gerado:")
    print("   - previsao_comparativo_2025.xlsx (relatório Excel com gráficos integrados)")

if __name__ == "__main__":
    main()
