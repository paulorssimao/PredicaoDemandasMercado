import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from .utils import safe_round, load_config


def calculate_inventory_levels(demand, lead_time=None, service_level=None):
    config = load_config()
    lead_time = lead_time or config['lead_time_days']
    service_level = service_level or config['service_level']
    avg_demand = demand.mean() if isinstance(demand, pd.Series) else demand
    std_demand = demand.std() if isinstance(demand, pd.Series) else demand * 0.3
    avg_demand = 0 if pd.isna(avg_demand) or pd.isnull(avg_demand) else avg_demand
    std_demand = 0 if pd.isna(std_demand) or pd.isnull(std_demand) else std_demand
    z = 1.645
    safety_stock = z * (lead_time ** 0.5) * std_demand
    estoque_minimo = (avg_demand * lead_time / 30) + safety_stock
    estoque_ideal = (avg_demand * (lead_time / 30 + 1)) + safety_stock
    estoque_maximo = estoque_ideal + (avg_demand * 0.5)
    return {'estoque_minimo': max(estoque_minimo, 0), 'estoque_ideal': max(estoque_ideal, 0), 'estoque_maximo': max(estoque_maximo, 0)}

def generate_excel_report(df_forecasts, output_file="analise_estoque_completo.xlsx"):
    wb = Workbook()

    def ajustar_largura_colunas(ws):
        for col in ws.columns:
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 10)

    # 1. Análise Mensal
    ws_monthly = wb.create_sheet("Análise Mensal")
    df_forecasts["mes"] = df_forecasts["data"].dt.to_period("M").dt.to_timestamp()
    monthly_grouped = df_forecasts.groupby(["mes", "CD_PRODUTO", "DS_PRODUTO", "UND_MED"])["qtd_prevista"]

    inventory_data = []
    for (mes, produto, nome, unidade), group in monthly_grouped:
        levels = calculate_inventory_levels(group)
        inventory_data.append({
            'Mês': mes,
            'Código Produto': produto,
            'Produto': nome,
            'Unidade': unidade,
            'Demanda Prevista': safe_round(group.sum()),
            'Estoque Mínimo': safe_round(levels['estoque_minimo']),
            'Estoque Ideal': safe_round(levels['estoque_ideal']),
            'Estoque Máximo': safe_round(levels['estoque_maximo'])
        })

    df_monthly = pd.DataFrame(inventory_data)
    for r in dataframe_to_rows(df_monthly, index=False, header=True):
        ws_monthly.append(r)
    ajustar_largura_colunas(ws_monthly)

    # 2. Análise Semanal
    ws_weekly = wb.create_sheet("Análise Semanal")
    df_forecasts["semana"] = df_forecasts["data"].dt.to_period("W").dt.to_timestamp()
    weekly_grouped = df_forecasts.groupby(["semana", "CD_PRODUTO", "DS_PRODUTO", "UND_MED"])["qtd_prevista"]

    inventory_data = []
    for (semana, produto, nome, unidade), group in weekly_grouped:
        levels = calculate_inventory_levels(group, lead_time=7)
        inventory_data.append({
            'Semana': semana,
            'Código Produto': produto,
            'Produto': nome,
            'Unidade': unidade,
            'Demanda Prevista': safe_round(group.sum()),
            'Estoque Mínimo': safe_round(levels['estoque_minimo']),
            'Estoque Ideal': safe_round(levels['estoque_ideal']),
            'Estoque Máximo': safe_round(levels['estoque_maximo'])
        })

    df_weekly = pd.DataFrame(inventory_data)
    for r in dataframe_to_rows(df_weekly, index=False, header=True):
        ws_weekly.append(r)
    ajustar_largura_colunas(ws_weekly)

    # 3. Resumo
    ws_summary = wb.create_sheet("Resumo")
    produtos = df_forecasts["CD_PRODUTO"].unique()

    resumo_data = []
    for produto in produtos:
        df_prod = df_forecasts[df_forecasts["CD_PRODUTO"] == produto]
        nome = df_prod["DS_PRODUTO"].iloc[0]
        unidade = df_prod["UND_MED"].iloc[0]

        monthly = df_prod.resample("M", on="data")["qtd_prevista"].sum()
        weekly = df_prod.resample("W", on="data")["qtd_prevista"].sum()

        m = calculate_inventory_levels(monthly)
        w = calculate_inventory_levels(weekly, lead_time=7)

        resumo_data.append({
            "Código Produto": produto,
            "Produto": nome,
            "Unidade": unidade,
            "Demanda Média Mensal": safe_round(monthly.mean()),
            "Estoque Mín Mensal": safe_round(m['estoque_minimo']),
            "Estoque Ideal Mensal": safe_round(m['estoque_ideal']),
            "Estoque Max Mensal": safe_round(m['estoque_maximo']),
            "Demanda Média Semanal": safe_round(weekly.mean()),
            "Estoque Mín Semanal": safe_round(w['estoque_minimo']),
            "Estoque Ideal Semanal": safe_round(w['estoque_ideal']),
            "Estoque Max Semanal": safe_round(w['estoque_maximo']),
        })

    df_resumo = pd.DataFrame(resumo_data)
    for r in dataframe_to_rows(df_resumo, index=False, header=True):
        ws_summary.append(r)
    ajustar_largura_colunas(ws_summary)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_file)
    print(f"✅ Relatório gerado: {output_file}")